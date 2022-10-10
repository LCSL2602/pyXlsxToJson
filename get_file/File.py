import openpyxl
from pathlib import Path


class File:
    def __init__(self, name_file: str):
        self.name_file = name_file

    def get_sheet(self, sheet_name=None) -> openpyxl.workbook.Workbook.active:
        try:
            xls_file = Path(f'{self.name_file}')
            wb_obj: openpyxl.Workbook = openpyxl.load_workbook(xls_file)
        except Exception as e:
            raise e
        else:
            if sheet_name is None:
                sheet: openpyxl.workbook.Workbook.active = wb_obj.active
            else:
                sheet: openpyxl.workbook.Workbook.active = wb_obj[sheet_name]
            return sheet

    def get_header(self) -> list[str]:
        list_header: list[str] = []
        file = self.get_sheet()
        if file.max_column > 0:
            index = 1
            while index <= file.max_column:
                list_header.append(file.cell(1, index).value)
                index += 1
        return list_header
